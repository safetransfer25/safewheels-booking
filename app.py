import io
import json
import os
import re
import shutil
import secrets
import sqlite3
import threading
from datetime import date, datetime, timedelta
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, unquote, urlparse
from zoneinfo import ZoneInfo

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

ROOT = Path(__file__).parent
PUBLIC = ROOT / "public"
DB_PATH = Path(os.environ.get("SAFEWHEELS_DB", ROOT / "safewheels.sqlite"))
BACKUP_DIR = Path(os.environ.get("SAFEWHEELS_BACKUP_DIR", DB_PATH.parent / "backups"))
LIVE_BACKUP_DIR = Path(os.environ.get("SAFEWHEELS_LIVE_BACKUP_DIR", BACKUP_DIR))
MAX_BACKUPS = 20
DB_LOCK = threading.RLock()
BOOKINGS_SCHEMA_COLUMNS = {
    "id",
    "date",
    "pickupTime",
    "travelTime",
    "flightNumber",
    "customerName",
    "phone",
    "hotel",
    "route",
    "passengers",
    "luggage",
    "price",
    "deposit",
    "balance",
    "paymentStatus",
    "paymentMethod",
    "vehicle",
    "driver",
    "bookingSource",
    "taxStatus",
    "status",
    "notes",
    "createdAt",
    "updatedAt",
}
XLSX_EXPORT_COLUMNS = [
    ("ID", "id"),
    ("Created At", "createdAt"),
    ("Updated At", "updatedAt"),
    ("Date", "date"),
    ("Time", "pickupTime"),
    ("Client", "customerName"),
    ("Phone", "phone"),
    ("Pickup", "hotel"),
    ("Dropoff", "route"),
    ("Flight", "flightNumber"),
    ("Travel Time", "travelTime"),
    ("Passengers", "passengers"),
    ("Luggage", "luggage"),
    ("Price", "price"),
    ("Deposit", "deposit"),
    ("Balance", "balance"),
    ("Payment Status", "paymentStatus"),
    ("Payment Method", "paymentMethod"),
    ("Driver", "driver"),
    ("Vehicle", "vehicle"),
    ("Booking Source", "bookingSource"),
    ("Status", "status"),
    ("AADE Status", "taxStatus"),
    ("Notes", "notes"),
]
PORT = int(os.environ.get("PORT", "4173"))
SESSIONS = set()
LOGIN_USERNAME = os.environ.get("SAFEWHEELS_USERNAME", "admin")
LOGIN_PASSWORD = os.environ.get("SAFEWHEELS_PASSWORD", "Safe2026&&&")
SEED_DEMO_DATA = os.environ.get("SAFEWHEELS_SEED_DEMO", "1") == "1"
APP_TZ = ZoneInfo(os.environ.get("SAFEWHEELS_TZ", "Europe/Athens"))
SEASON_MONTHS = [
    (4, "Απρίλιος"),
    (5, "Μάιος"),
    (6, "Ιούνιος"),
    (7, "Ιούλιος"),
    (8, "Αύγουστος"),
    (9, "Σεπτέμβριος"),
    (10, "Οκτώβριος"),
    (11, "Νοέμβριος"),
]

DEFAULT_VEHICLES = ["OPEL VIVARO", "PEUGEOT 5008"]
DEFAULT_DRIVERS = ["Θεόδωρος Τσιάμης", "Γεώργιος Τσιάμης", "Ιωάννης Τσιάμης"]
DEFAULT_BOOKING_SOURCES = ["PRIVATE", "WELCOME", "CONNECTO"]
PAYMENT_METHODS = {"Μετρητά", "Κάρτα", "Πίστωση"}
TAX_STATUSES = {"Καταχωρημένο", "Μη Καταχωρημένο"}
BOOKING_STATUSES = {"Pending", "Confirmed", "Completed", "Cancelled"}


def db():
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def table_columns(conn, table):
    rows = conn.execute(f"PRAGMA table_info({table})").fetchall()
    return {row[1] for row in rows}


def ensure_option_table(conn, table):
    columns = table_columns(conn, table)
    if "active" not in columns:
        conn.execute(f"ALTER TABLE {table} ADD COLUMN active INTEGER DEFAULT 1")
    if "createdAt" not in columns:
        conn.execute(f"ALTER TABLE {table} ADD COLUMN createdAt TEXT")
        conn.execute(f"UPDATE {table} SET createdAt = CURRENT_TIMESTAMP WHERE createdAt IS NULL")
    if "updatedAt" not in columns:
        conn.execute(f"ALTER TABLE {table} ADD COLUMN updatedAt TEXT")
        conn.execute(f"UPDATE {table} SET updatedAt = CURRENT_TIMESTAMP WHERE updatedAt IS NULL")


def ensure_booking_audit_table(conn):
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS booking_audit_log (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          bookingId INTEGER NOT NULL,
          action TEXT NOT NULL,
          loggedAt TEXT DEFAULT CURRENT_TIMESTAMP,
          createdAt TEXT,
          updatedAt TEXT,
          status TEXT,
          driver TEXT,
          vehicle TEXT,
          paymentMethod TEXT,
          taxStatus TEXT,
          snapshot TEXT
        )
        """
    )


def upsert_option(conn, table, name):
    clean = str(name or "").strip()
    if not clean:
        return
    existing = conn.execute(
        f"SELECT rowid AS option_rowid FROM {table} WHERE name = ? COLLATE NOCASE LIMIT 1",
        (clean,),
    ).fetchone()
    if existing:
        conn.execute(
            f"UPDATE {table} SET name = ?, active = 1, updatedAt = CURRENT_TIMESTAMP WHERE rowid = ?",
            (clean, existing["option_rowid"]),
        )
        return
    conn.execute(f"INSERT INTO {table} (name, active) VALUES (?, 1)", (clean,))


def seed_options(conn):
    for name in DEFAULT_VEHICLES:
        upsert_option(conn, "vehicles", name)
    for name in DEFAULT_DRIVERS:
        upsert_option(conn, "drivers", name)
    for name in DEFAULT_BOOKING_SOURCES:
        upsert_option(conn, "booking_sources", name)
    for row in conn.execute("SELECT DISTINCT vehicle AS name FROM bookings WHERE vehicle IS NOT NULL AND vehicle != ''").fetchall():
        upsert_option(conn, "vehicles", row["name"])
    for row in conn.execute("SELECT DISTINCT driver AS name FROM bookings WHERE driver IS NOT NULL AND driver != ''").fetchall():
        upsert_option(conn, "drivers", row["name"])
    for row in conn.execute("SELECT DISTINCT bookingSource AS name FROM bookings WHERE bookingSource IS NOT NULL AND bookingSource != ''").fetchall():
        upsert_option(conn, "booking_sources", row["name"])


def active_option_names(conn, table):
    return [
        row["name"]
        for row in conn.execute(f"SELECT name FROM {table} WHERE active = 1 ORDER BY name COLLATE NOCASE").fetchall()
    ]


def active_option_rows(conn, table):
    return [
        row_to_dict(row)
        for row in conn.execute(f"SELECT id, name FROM {table} WHERE active = 1 ORDER BY name COLLATE NOCASE").fetchall()
    ]


def init_db():
    backup_database("schema-change")
    with db() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS bookings (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              date TEXT NOT NULL,
              pickupTime TEXT NOT NULL,
              travelTime TEXT,
              flightNumber TEXT,
              customerName TEXT NOT NULL,
              phone TEXT,
              hotel TEXT,
              route TEXT,
              passengers INTEGER DEFAULT 1,
              luggage INTEGER DEFAULT 0,
              price REAL DEFAULT 0,
              deposit REAL DEFAULT 0,
              balance REAL DEFAULT 0,
              paymentStatus TEXT DEFAULT 'Unpaid',
              paymentMethod TEXT DEFAULT 'Μετρητά',
              vehicle TEXT DEFAULT 'OPEL VIVARO',
              driver TEXT,
              bookingSource TEXT DEFAULT 'PRIVATE',
              taxStatus TEXT DEFAULT 'Μη Καταχωρημένο',
              status TEXT DEFAULT 'Pending',
              notes TEXT,
              createdAt TEXT DEFAULT CURRENT_TIMESTAMP,
              updatedAt TEXT DEFAULT CURRENT_TIMESTAMP
            )
            """
        )
        columns = table_columns(conn, "bookings")
        if "paymentMethod" not in columns:
            conn.execute("ALTER TABLE bookings ADD COLUMN paymentMethod TEXT DEFAULT 'Μετρητά'")
        if "flightNumber" not in columns:
            conn.execute("ALTER TABLE bookings ADD COLUMN flightNumber TEXT")
        if "bookingSource" not in columns:
            conn.execute("ALTER TABLE bookings ADD COLUMN bookingSource TEXT DEFAULT 'PRIVATE'")
        if "taxStatus" not in columns:
            conn.execute("ALTER TABLE bookings ADD COLUMN taxStatus TEXT DEFAULT 'Μη Καταχωρημένο'")
        conn.execute("UPDATE bookings SET vehicle = 'OPEL VIVARO' WHERE vehicle IN ('SafeWheels 1', 'SafeWheels1')")
        conn.execute("UPDATE bookings SET vehicle = 'PEUGEOT 5008' WHERE vehicle IN ('SafeWheels 2', 'SafeWheels2')")
        conn.execute("UPDATE bookings SET paymentMethod = 'Μετρητά' WHERE paymentMethod IS NULL OR paymentMethod = ''")
        conn.execute("UPDATE bookings SET bookingSource = 'PRIVATE' WHERE bookingSource IS NULL OR bookingSource = ''")
        conn.execute("UPDATE bookings SET taxStatus = 'Μη Καταχωρημένο' WHERE taxStatus IS NULL OR taxStatus = ''")
        repair_booking_dates(conn)

        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS expenses (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              date TEXT NOT NULL,
              category TEXT NOT NULL DEFAULT 'Καύσιμα',
              description TEXT,
              amount REAL DEFAULT 0,
              paymentMethod TEXT DEFAULT 'Μετρητά',
              vehicle TEXT,
              notes TEXT,
              createdAt TEXT DEFAULT CURRENT_TIMESTAMP,
              updatedAt TEXT DEFAULT CURRENT_TIMESTAMP
            )
            """
        )
        expense_columns = table_columns(conn, "expenses")
        if "vehicle" not in expense_columns:
            conn.execute("ALTER TABLE expenses ADD COLUMN vehicle TEXT")

        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS vehicles (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              name TEXT NOT NULL UNIQUE,
              active INTEGER DEFAULT 1,
              createdAt TEXT DEFAULT CURRENT_TIMESTAMP,
              updatedAt TEXT DEFAULT CURRENT_TIMESTAMP
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS drivers (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              name TEXT NOT NULL UNIQUE,
              active INTEGER DEFAULT 1,
              createdAt TEXT DEFAULT CURRENT_TIMESTAMP,
              updatedAt TEXT DEFAULT CURRENT_TIMESTAMP
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS booking_sources (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              name TEXT NOT NULL UNIQUE,
              active INTEGER DEFAULT 1,
              createdAt TEXT DEFAULT CURRENT_TIMESTAMP,
              updatedAt TEXT DEFAULT CURRENT_TIMESTAMP
            )
            """
        )
        ensure_booking_audit_table(conn)
        ensure_option_table(conn, "vehicles")
        ensure_option_table(conn, "drivers")
        ensure_option_table(conn, "booking_sources")
        seed_options(conn)
        ensure_booking_audit_baseline(conn)

        total = conn.execute("SELECT COUNT(*) AS total FROM bookings").fetchone()["total"]
        if total or not SEED_DEMO_DATA:
            return
        today = datetime.now().date().isoformat()
        seed = [
            (today, "08:20", "10:05", "A3 1234", "Maria Jensen", "+45 20 12 45 88", "Aqua Blu Boutique Hotel", "Αεροδρόμιο Κω -> Ξενοδοχείο", 2, 2, 45, "Paid", "Μετρητά", "OPEL VIVARO", "Θεόδωρος Τσιάμης", "PRIVATE", "Μη Καταχωρημένο", "Confirmed", "Παιδικό κάθισμα"),
            (today, "14:10", "15:00", "FR 2451", "Luca Moretti", "+39 333 700 1200", "Kos Aktis Art Hotel", "Λιμάνι Κω -> Ξενοδοχείο", 4, 3, 35, "Unpaid", "Κάρτα", "PEUGEOT 5008", "Γεώργιος Τσιάμης", "WELCOME", "Μη Καταχωρημένο", "Pending", "Άφιξη με ferry από Ρόδο"),
        ]
        conn.executemany(
            """
            INSERT INTO bookings
            (date, pickupTime, travelTime, flightNumber, customerName, phone, hotel, route, passengers, luggage, price, paymentStatus, paymentMethod, vehicle, driver, bookingSource, taxStatus, status, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            seed,
        )


def row_to_dict(row):
    return dict(row) if row else None


def utc_timestamp():
    return datetime.utcnow().replace(microsecond=0).isoformat() + "Z"


def backup_database(reason):
    if not DB_PATH.exists():
        return None
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now(APP_TZ).strftime("%Y_%m_%d_%H%M")
    backup_path = BACKUP_DIR / f"safewheels_backup_{timestamp}.db"
    suffix = 1
    while backup_path.exists():
        backup_path = BACKUP_DIR / f"safewheels_backup_{timestamp}_{suffix:02d}.db"
        suffix += 1
    shutil.copy2(DB_PATH, backup_path)
    prune_backups()
    return backup_path


def prune_backups():
    backups = sorted(
        BACKUP_DIR.glob("safewheels_backup_*.db"),
        key=lambda item: item.stat().st_mtime,
        reverse=True,
    )
    for old_backup in backups[MAX_BACKUPS:]:
        old_backup.unlink(missing_ok=True)


def live_backup_timestamp():
    return datetime.now(APP_TZ).strftime("%Y-%m-%d_%H-%M-%S")


def is_allowed_live_backup_name(name):
    clean = Path(unquote(name)).name
    if clean != name or ".." in name or "/" in name or "\\" in name:
        return False
    if not clean.endswith(".sqlite"):
        return False
    return clean.startswith("safewheels_live_backup_") or clean.startswith(
        "safewheels_emergency_before_restore_"
    )


def copy_sqlite_file(source, destination):
    destination.parent.mkdir(parents=True, exist_ok=True)
    src = sqlite3.connect(f"file:{source}?mode=ro", uri=True)
    dst = sqlite3.connect(destination)
    try:
        src.backup(dst)
        dst.commit()
    finally:
        dst.close()
        src.close()


def create_live_backup(kind="live"):
    if not DB_PATH.exists():
        raise FileNotFoundError("Database file not found.")
    LIVE_BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = live_backup_timestamp()
    prefix = (
        "safewheels_emergency_before_restore_"
        if kind == "emergency"
        else "safewheels_live_backup_"
    )
    backup_path = LIVE_BACKUP_DIR / f"{prefix}{timestamp}.sqlite"
    suffix = 1
    while backup_path.exists():
        backup_path = LIVE_BACKUP_DIR / f"{prefix}{timestamp}_{suffix:02d}.sqlite"
        suffix += 1
    with DB_LOCK:
        copy_sqlite_file(DB_PATH, backup_path)
    return backup_path


def list_live_backups():
    LIVE_BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    items = []
    for path in LIVE_BACKUP_DIR.glob("safewheels_*.sqlite"):
        if not path.is_file():
            continue
        stat = path.stat()
        items.append(
            {
                "filename": path.name,
                "size": stat.st_size,
                "createdAt": datetime.fromtimestamp(stat.st_mtime, APP_TZ).isoformat(),
            }
        )
    items.sort(key=lambda item: item["createdAt"], reverse=True)
    return items


def validate_sqlite_backup(path):
    if not path.exists() or path.stat().st_size < 100:
        return False, "Το αρχείο SQLite είναι κενό ή πολύ μικρό."
    header = path.read_bytes()[:16]
    if not header.startswith(b"SQLite format 3"):
        return False, "Το αρχείο δεν είναι έγκυρο SQLite database."
    try:
        conn = sqlite3.connect(f"file:{path}?mode=ro", uri=True)
        tables = {
            row[0]
            for row in conn.execute(
                "SELECT name FROM sqlite_master WHERE type='table'"
            ).fetchall()
        }
        if "bookings" not in tables:
            conn.close()
            return False, "Λείπει ο πίνακας bookings."
        columns = table_columns(conn, "bookings")
        missing = sorted(BOOKINGS_SCHEMA_COLUMNS - columns)
        if missing:
            conn.close()
            return False, f"Ασύμβατο schema bookings. Λείπουν: {', '.join(missing)}"
        count = conn.execute("SELECT COUNT(*) FROM bookings").fetchone()[0]
        conn.execute("SELECT id, date, customerName FROM bookings LIMIT 1").fetchone()
        conn.close()
        return True, {"bookingCount": count, "columns": sorted(columns)}
    except sqlite3.Error as exc:
        return False, f"Σφάλμα ανάγνωσης SQLite: {exc}"


def restore_sqlite_backup(upload_path):
    with DB_LOCK:
        emergency_path = create_live_backup("emergency")
        valid, details = validate_sqlite_backup(upload_path)
        if not valid:
            return False, details, emergency_path.name
        if DB_PATH.exists():
            copy_sqlite_file(upload_path, DB_PATH)
        else:
            DB_PATH.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(upload_path, DB_PATH)
        verify_conn = sqlite3.connect(f"file:{DB_PATH}?mode=ro", uri=True)
        verify_conn.row_factory = sqlite3.Row
        try:
            count = verify_conn.execute("SELECT COUNT(*) FROM bookings").fetchone()[0]
            verify_conn.execute("SELECT id FROM bookings LIMIT 1").fetchone()
        finally:
            verify_conn.close()
        return True, {"bookingCount": count, "emergencyBackup": emergency_path.name}, emergency_path.name


def build_bookings_xlsx():
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl is not installed.")
    with db() as conn:
        rows = conn.execute("SELECT * FROM bookings ORDER BY id ASC").fetchall()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Bookings"
    headers = [label for label, _ in XLSX_EXPORT_COLUMNS]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for row in rows:
        data = row_to_dict(row)
        sheet.append([data.get(field, "") for _, field in XLSX_EXPORT_COLUMNS])
    sheet.freeze_panes = "A2"
    for index, (label, _) in enumerate(XLSX_EXPORT_COLUMNS, start=1):
        values = [len(str(label))]
        for excel_row in sheet.iter_rows(
            min_row=2, min_col=index, max_col=index, values_only=True
        ):
            values.append(len(str(excel_row[0] or "")))
        sheet.column_dimensions[get_column_letter(index)].width = min(max(values) + 2, 48)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def parse_multipart_upload(content_type, body):
    boundary = None
    for piece in content_type.split(";"):
        piece = piece.strip()
        if piece.startswith("boundary="):
            boundary = piece.split("=", 1)[1].strip().strip('"')
    if not boundary:
        raise ValueError("Missing multipart boundary.")
    delimiter = ("--" + boundary).encode()
    for part in body.split(delimiter):
        if b"filename=" not in part:
            continue
        _, _, payload = part.partition(b"\r\n\r\n")
        if not payload:
            continue
        if payload.endswith(b"\r\n"):
            payload = payload[:-2]
        if payload.endswith(b"--"):
            payload = payload[:-2]
        if payload.startswith(b"\r\n"):
            payload = payload[2:]
        return payload
    raise ValueError("Missing uploaded file.")


def audit_booking(conn, booking_id, action):
    ensure_booking_audit_table(conn)
    row = conn.execute("SELECT * FROM bookings WHERE id = ?", (booking_id,)).fetchone()
    if not row:
        return
    conn.execute(
        """
        INSERT INTO booking_audit_log
        (bookingId, action, createdAt, updatedAt, status, driver, vehicle, paymentMethod, taxStatus, snapshot)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            row["id"],
            action,
            row["createdAt"],
            row["updatedAt"],
            row["status"],
            row["driver"],
            row["vehicle"],
            row["paymentMethod"],
            row["taxStatus"],
            json.dumps(row_to_dict(row), ensure_ascii=False),
        ),
    )


def persist_booking_audit(booking_id, action):
    try:
        with db() as conn:
            audit_booking(conn, booking_id, action)
    except Exception:
        # Audit runs after commit; never undo a saved booking.
        pass


def ensure_booking_audit_baseline(conn):
    ensure_booking_audit_table(conn)
    rows = conn.execute(
        """
        SELECT id
        FROM bookings
        WHERE id NOT IN (SELECT DISTINCT bookingId FROM booking_audit_log)
        """
    ).fetchall()
    for row in rows:
        audit_booking(conn, row["id"], "baseline")


def parse_date(value):
    return datetime.strptime(value, "%Y-%m-%d").date()


def app_today():
    return datetime.now(APP_TZ).date()


def normalize_date(value):
    raw = str(value or "").strip()
    if not raw:
        return ""
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", raw):
        try:
            return datetime.strptime(raw, "%Y-%m-%d").date().isoformat()
        except ValueError:
            return ""
    match = re.fullmatch(r"(\d{1,2})[/-](\d{1,2})(?:[/-](\d{2}|\d{4}))?", raw)
    if match:
        day = int(match.group(1))
        month = int(match.group(2))
        year_value = match.group(3)
        year = app_today().year if not year_value else int(year_value)
        if year < 100:
            year += 2000
        try:
            return date(year, month, day).isoformat()
        except ValueError:
            return ""
    return ""


def normalize_time(value):
    raw = str(value or "").strip()
    if not raw:
        return ""
    match = re.fullmatch(r"([01]?\d|2[0-3]):([0-5]\d)", raw)
    if match:
        return f"{int(match.group(1)):02d}:{match.group(2)}"
    match = re.fullmatch(r"(\d{1,2})([0-5]\d)", raw)
    if match and int(match.group(1)) <= 23:
        return f"{int(match.group(1)):02d}:{match.group(2)}"
    return raw


def normalize_number(value, default=0):
    raw = str(value if value is not None else "").strip().replace(",", ".")
    if not raw:
        return default
    try:
        return float(raw)
    except ValueError:
        return default


def normalize_int(value, default=0):
    try:
        return int(normalize_number(value, default))
    except (TypeError, ValueError):
        return default


def repair_booking_dates(conn):
    rows = conn.execute("SELECT id, date, pickupTime, travelTime, status FROM bookings").fetchall()
    for row in rows:
        fixed_date = normalize_date(row["date"])
        fixed_pickup = normalize_time(row["pickupTime"])
        fixed_travel = normalize_time(row["travelTime"])
        fixed_status = row["status"] if row["status"] in BOOKING_STATUSES else "Pending"
        updates = {}
        if fixed_date and fixed_date != row["date"]:
            updates["date"] = fixed_date
        if fixed_pickup and fixed_pickup != row["pickupTime"]:
            updates["pickupTime"] = fixed_pickup
        if fixed_travel != (row["travelTime"] or ""):
            updates["travelTime"] = fixed_travel
        if fixed_status != row["status"]:
            updates["status"] = fixed_status
        if updates:
            assignments = ", ".join([f"{key}=:{key}" for key in updates])
            conn.execute(
                f"UPDATE bookings SET {assignments}, updatedAt=CURRENT_TIMESTAMP WHERE id=:id",
                {**updates, "id": row["id"]},
            )


def month_bounds(month):
    start = datetime.strptime(month, "%Y-%m").date().replace(day=1)
    next_month = (start.replace(day=28) + timedelta(days=4)).replace(day=1)
    return start.isoformat(), (next_month - timedelta(days=1)).isoformat()


def parse_report_period(query):
    now = datetime.now(APP_TZ)
    try:
        year = int(query.get("year", [str(now.year)])[0])
    except (TypeError, ValueError):
        year = now.year
    try:
        month = int(query.get("month", [str(now.month)])[0])
    except (TypeError, ValueError):
        month = now.month
    month = max(1, min(12, month))
    month_key = f"{year:04d}-{month:02d}"
    start, end = month_bounds(month_key)
    return year, month, month_key, start, end


def fetch_monthly_report(year, month):
    _, _, month_key, start, end = parse_report_period({"year": [str(year)], "month": [str(month)]})
    with db() as conn:
        summary = conn.execute(
            """
            SELECT
              COUNT(*) AS totalTransfers,
              COALESCE(SUM(price), 0) AS totalRevenue,
              COALESCE(SUM(CASE WHEN paymentMethod = 'Μετρητά' THEN price ELSE 0 END), 0) AS cashRevenue,
              COALESCE(SUM(CASE WHEN paymentMethod = 'Κάρτα' THEN price ELSE 0 END), 0) AS cardRevenue,
              COALESCE(SUM(CASE WHEN paymentMethod = 'Πίστωση' THEN price ELSE 0 END), 0) AS creditRevenue,
              COALESCE(SUM(CASE WHEN paymentMethod NOT IN ('Μετρητά', 'Κάρτα', 'Πίστωση') THEN price ELSE 0 END), 0) AS bankTransferRevenue,
              COALESCE(SUM(CASE WHEN UPPER(COALESCE(bookingSource, 'PRIVATE')) = 'PRIVATE' THEN price ELSE 0 END), 0) AS privateRevenue,
              COALESCE(SUM(CASE WHEN UPPER(COALESCE(bookingSource, 'PRIVATE')) = 'WELCOME' THEN price ELSE 0 END), 0) AS welcomeRevenue,
              COALESCE(SUM(CASE WHEN UPPER(COALESCE(bookingSource, 'PRIVATE')) IN ('GETTRANSFER', 'CONNECTO') THEN price ELSE 0 END), 0) AS gettransferRevenue
            FROM bookings
            WHERE date BETWEEN ? AND ?
              AND status != 'Cancelled'
            """,
            (start, end),
        ).fetchone()
        expenses = conn.execute(
            "SELECT COALESCE(SUM(amount), 0) AS total FROM expenses WHERE date BETWEEN ? AND ?",
            (start, end),
        ).fetchone()["total"]
        by_source = [
            row_to_dict(row)
            for row in conn.execute(
                """
                SELECT
                  UPPER(COALESCE(bookingSource, 'PRIVATE')) AS source,
                  COUNT(*) AS transfers,
                  COALESCE(SUM(price), 0) AS revenue
                FROM bookings
                WHERE date BETWEEN ? AND ?
                  AND status != 'Cancelled'
                GROUP BY UPPER(COALESCE(bookingSource, 'PRIVATE'))
                ORDER BY revenue DESC, source ASC
                """,
                (start, end),
            ).fetchall()
        ]
    data = row_to_dict(summary)
    total_revenue = float(data["totalRevenue"] or 0)
    expense_total = float(expenses or 0)
    card_total = float(data["cardRevenue"] or 0) + float(data["creditRevenue"] or 0)
    return {
        "year": year,
        "month": month,
        "monthKey": month_key,
        "periodStart": start,
        "periodEnd": end,
        "totalTransfers": int(data["totalTransfers"] or 0),
        "totalRevenue": total_revenue,
        "cashRevenue": float(data["cashRevenue"] or 0),
        "cardRevenue": card_total,
        "cardOnlyRevenue": float(data["cardRevenue"] or 0),
        "creditRevenue": float(data["creditRevenue"] or 0),
        "bankTransferRevenue": float(data["bankTransferRevenue"] or 0),
        "privateRevenue": float(data["privateRevenue"] or 0),
        "welcomeRevenue": float(data["welcomeRevenue"] or 0),
        "gettransferRevenue": float(data["gettransferRevenue"] or 0),
        "expenses": expense_total,
        "profit": total_revenue - expense_total,
        "rowsUsedCount": int(data["totalTransfers"] or 0),
        "bySource": by_source,
    }


def build_monthly_report_xlsx(report):
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl is not installed.")
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary.append(["Monthly Report", f"{report['monthKey']}"])
    summary.append([])
    rows = [
        ("Total Transfers", report["totalTransfers"]),
        ("Total Revenue", report["totalRevenue"]),
        ("Cash", report["cashRevenue"]),
        ("Card", report["cardRevenue"]),
        ("PRIVATE Revenue", report["privateRevenue"]),
        ("WELCOME Revenue", report["welcomeRevenue"]),
        ("GETTRANSFER Revenue", report["gettransferRevenue"]),
        ("Expenses", report["expenses"]),
        ("Profit", report["profit"]),
        ("Rows Used", report["rowsUsedCount"]),
    ]
    if report["bankTransferRevenue"]:
        rows.insert(5, ("Bank Transfer / Other", report["bankTransferRevenue"]))
    for label, value in rows:
        summary.append([label, value])
    summary.freeze_panes = "A3"
    summary["A1"].font = Font(bold=True)
    sources = workbook.create_sheet("By Source")
    sources.append(["Source", "Transfers", "Revenue"])
    for cell in sources[1]:
        cell.font = Font(bold=True)
    for row in report["bySource"]:
        sources.append([row["source"], row["transfers"], row["revenue"]])
    sources.freeze_panes = "A2"
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def normalize_booking(data):
    price = normalize_number(data.get("price"), 0)
    vehicle = str(data.get("vehicle") or "OPEL VIVARO")
    driver = str(data.get("driver") or "").strip()
    payment_method = str(data.get("paymentMethod") or "Μετρητά")
    booking_source = str(data.get("bookingSource") or "PRIVATE").strip().upper()
    tax_status = str(data.get("taxStatus") or "Μη Καταχωρημένο").strip()
    status = str(data.get("status") or "Pending").strip()
    return {
        "date": normalize_date(data.get("date")),
        "pickupTime": normalize_time(data.get("pickupTime")),
        "travelTime": normalize_time(data.get("travelTime")),
        "flightNumber": str(data.get("flightNumber") or "").strip().upper(),
        "customerName": str(data.get("customerName") or "").strip(),
        "phone": str(data.get("phone") or "").strip(),
        "hotel": str(data.get("hotel") or "").strip(),
        "route": str(data.get("route") or "").strip(),
        "passengers": normalize_int(data.get("passengers"), 1),
        "luggage": normalize_int(data.get("luggage"), 0),
        "price": price,
        "paymentStatus": str(data.get("paymentStatus") or "Unpaid"),
        "paymentMethod": payment_method if payment_method in PAYMENT_METHODS else "Μετρητά",
        "vehicle": vehicle,
        "driver": driver,
        "bookingSource": booking_source or "PRIVATE",
        "taxStatus": tax_status if tax_status in TAX_STATUSES else "Μη Καταχωρημένο",
        "status": status if status in BOOKING_STATUSES else "Pending",
        "notes": str(data.get("notes") or "").strip(),
    }


def normalize_expense(data):
    payment_method = str(data.get("paymentMethod") or "Μετρητά")
    return {
        "date": normalize_date(data.get("date")),
        "category": str(data.get("category") or "Καύσιμα").strip(),
        "description": str(data.get("description") or "").strip(),
        "amount": normalize_number(data.get("amount"), 0),
        "paymentMethod": payment_method if payment_method in PAYMENT_METHODS else "Μετρητά",
        "vehicle": str(data.get("vehicle") or "").strip(),
        "notes": str(data.get("notes") or "").strip(),
    }


def booking_error(booking):
    if not booking["date"]:
        return "Η ημερομηνία είναι υποχρεωτική και πρέπει να είναι σε μορφή YYYY-MM-DD ή ΗΗ/ΜΜ/ΕΕΕΕ."
    if not booking["pickupTime"]:
        return "Η ώρα παραλαβής είναι υποχρεωτική."
    if not booking["customerName"]:
        return "Το όνομα πελάτη είναι υποχρεωτικό."
    return None


def expense_error(expense):
    if not expense["date"]:
        return "Η ημερομηνία είναι υποχρεωτική."
    if not expense["category"]:
        return "Η κατηγορία εξόδου είναι υποχρεωτική."
    if expense["amount"] < 0:
        return "Το ποσό δεν μπορεί να είναι αρνητικό."
    return None


class Handler(SimpleHTTPRequestHandler):
    def end_headers(self):
        self.send_header("Cache-Control", "no-store, no-cache, must-revalidate, max-age=0")
        self.send_header("Pragma", "no-cache")
        self.send_header("Expires", "0")
        super().end_headers()

    def translate_path(self, path):
        clean = urlparse(path).path.lstrip("/")
        if not clean:
            clean = "index.html"
        return str(PUBLIC / clean)

    def do_POST(self):
        if self.path == "/api/login":
            data = self.read_json()
            if data.get("username") == LOGIN_USERNAME and data.get("password") == LOGIN_PASSWORD:
                token = secrets.token_urlsafe(32)
                SESSIONS.add(token)
                return self.send_json({"token": token})
            return self.send_json({"error": "Λάθος στοιχεία σύνδεσης."}, 401)
        if self.path == "/api/bookings":
            return self.create_booking()
        if self.path == "/api/expenses":
            return self.create_expense()
        if self.path.startswith("/api/options/"):
            return self.create_option()
        parsed = urlparse(self.path)
        if parsed.path == "/api/admin/create-backup":
            return self.admin_create_backup()
        if parsed.path == "/api/admin/restore-backup":
            return self.admin_restore_backup()
        self.send_error(404)

    def do_GET(self):
        parsed = urlparse(self.path)
        if parsed.path == "/api/bookings":
            return self.list_bookings(parsed)
        if parsed.path == "/api/totals":
            return self.totals(parsed)
        if parsed.path == "/api/expenses":
            return self.list_expenses(parsed)
        if parsed.path == "/api/expense-summary":
            return self.expense_summary(parsed)
        if parsed.path == "/api/monthly-revenue":
            return self.monthly_revenue(parsed)
        if parsed.path == "/api/reports/monthly":
            return self.reports_monthly(parsed)
        if parsed.path == "/api/reports/monthly/export-xlsx":
            return self.reports_monthly_export_xlsx(parsed)
        if parsed.path == "/api/options":
            return self.options()
        if parsed.path == "/api/admin/debug":
            return self.admin_debug(parsed)
        if parsed.path == "/api/admin/backups":
            return self.admin_list_backups()
        if parsed.path == "/api/admin/export-bookings-xlsx":
            return self.admin_export_bookings_xlsx()
        if parsed.path.startswith("/api/admin/download-backup/"):
            filename = parsed.path.rsplit("/", 1)[-1]
            return self.admin_download_backup(filename)
        if parsed.path.startswith("/api/"):
            return self.send_error(404)
        return super().do_GET()

    def do_PUT(self):
        parsed = urlparse(self.path)
        if parsed.path.startswith("/api/bookings/"):
            if parsed.path.endswith("/tax-status"):
                return self.update_booking_tax_status(parsed)
            if parsed.path.endswith("/status"):
                return self.update_booking_status(parsed)
            return self.update_booking(parsed)
        if parsed.path.startswith("/api/expenses/"):
            return self.update_expense(parsed)
        if parsed.path.startswith("/api/options/"):
            return self.update_option(parsed)
        return self.send_error(404)

    def do_DELETE(self):
        parsed = urlparse(self.path)
        if parsed.path.startswith("/api/bookings/"):
            return self.delete_row("bookings", parsed)
        if parsed.path.startswith("/api/expenses/"):
            return self.delete_row("expenses", parsed)
        if parsed.path.startswith("/api/options/"):
            return self.delete_option(parsed)
        return self.send_error(404)

    def create_booking(self):
        if not self.authorized():
            return self.send_json({"error": "Unauthorized"}, 401)
        booking = normalize_booking(self.read_json())
        if booking["status"] == "Completed":
            booking["status"] = "Pending"
        error = booking_error(booking)
        if error:
            return self.send_json({"error": error}, 400)
        keys = ", ".join(booking.keys())
        placeholders = ", ".join([f":{key}" for key in booking])
        backup_database("booking-create")
        with db() as conn:
            cur = conn.execute(f"INSERT INTO bookings ({keys}) VALUES ({placeholders})", booking)
            saved = conn.execute("SELECT * FROM bookings WHERE id = ?", (cur.lastrowid,)).fetchone()
        saved_booking = row_to_dict(saved)
        if not saved_booking or not saved_booking.get("id"):
            return self.send_json({"error": "Η κράτηση δεν επιβεβαιώθηκε στη βάση."}, 500)
        persist_booking_audit(saved_booking["id"], "created")
        return self.send_json(saved_booking, 201)

    def list_bookings(self, parsed):
        if not self.authorized():
            return self.send_json({"error": "Unauthorized"}, 401)
        query = parse_qs(parsed.query)
        history = query.get("history", ["0"])[0] == "1"
        where = []
        params = {}
        has_date_filter = False
        if query.get("date", [""])[0]:
            where.append("date = :date")
            params["date"] = normalize_date(query["date"][0])
            has_date_filter = True
        if query.get("start", [""])[0] and query.get("end", [""])[0]:
            where.append("date BETWEEN :start AND :end")
            params["start"] = normalize_date(query["start"][0])
            params["end"] = normalize_date(query["end"][0])
            has_date_filter = True
        if query.get("q", [""])[0]:
            where.append("(CAST(id AS TEXT) LIKE :q OR customerName LIKE :q OR phone LIKE :q OR hotel LIKE :q OR date LIKE :q OR pickupTime LIKE :q OR flightNumber LIKE :q OR route LIKE :q)")
            params["q"] = f"%{query['q'][0]}%"
        if query.get("tax", [""])[0] in TAX_STATUSES:
            where.append("taxStatus = :tax")
            params["tax"] = query["tax"][0]
        if history:
            where.append("status = 'Completed'")
        else:
            where.append("status NOT IN ('Completed', 'Cancelled')")
            if not has_date_filter:
                where.append("date >= :today")
                params["today"] = app_today().isoformat()
        sql = "SELECT * FROM bookings"
        if where:
            sql += " WHERE " + " AND ".join(where)
        sql += " ORDER BY date DESC, pickupTime DESC" if history else " ORDER BY date ASC, pickupTime ASC"
        with db() as conn:
            rows = conn.execute(sql, params).fetchall()
        return self.send_json([row_to_dict(row) for row in rows])

    def update_booking(self, parsed):
        if not self.authorized():
            return self.send_json({"error": "Unauthorized"}, 401)
        booking_id = parsed.path.rsplit("/", 1)[-1]
        booking = normalize_booking(self.read_json())
        error = booking_error(booking)
        if error:
            return self.send_json({"error": error}, 400)
        assignments = ", ".join([f"{key}=:{key}" for key in booking])
        backup_database("booking-update")
        with db() as conn:
            existing = conn.execute("SELECT status FROM bookings WHERE id = ?", (booking_id,)).fetchone()
            if not existing:
                return self.send_json({"error": "Δεν βρέθηκε η κράτηση."}, 404)
            if booking["status"] == "Completed" and existing["status"] != "Completed":
                booking["status"] = existing["status"] if existing["status"] in BOOKING_STATUSES else "Pending"
            cur = conn.execute(
                f"UPDATE bookings SET {assignments}, updatedAt=CURRENT_TIMESTAMP WHERE id=:id",
                {**booking, "id": booking_id},
            )
            if cur.rowcount == 0:
                return self.send_json({"error": "Δεν βρέθηκε η κράτηση."}, 404)
            saved = conn.execute("SELECT * FROM bookings WHERE id = ?", (booking_id,)).fetchone()
        persist_booking_audit(booking_id, "updated")
        return self.send_json(row_to_dict(saved))

    def totals(self, parsed):
        if not self.authorized():
            return self.send_json({"error": "Unauthorized"}, 401)
        query = parse_qs(parsed.query)
        start = normalize_date(query.get("start", [""])[0])
        end = normalize_date(query.get("end", [""])[0])
        with db() as conn:
            row = conn.execute(
                """
                SELECT
                  COUNT(*) AS bookings,
                  COALESCE(SUM(CASE WHEN paymentMethod = 'Μετρητά' THEN price ELSE 0 END), 0) AS cash,
                  COALESCE(SUM(CASE WHEN paymentMethod IN ('Κάρτα', 'Πίστωση') THEN price ELSE 0 END), 0) AS card,
                  COALESCE(SUM(price), 0) AS total
                FROM bookings
                WHERE date BETWEEN ? AND ?
                  AND status NOT IN ('Completed', 'Cancelled')
                """,
                (start, end),
            ).fetchone()
        return self.send_json(row_to_dict(row))

    def create_expense(self):
        if not self.authorized():
            return self.send_json({"error": "Unauthorized"}, 401)
        expense = normalize_expense(self.read_json())
        error = expense_error(expense)
        if error:
            return self.send_json({"error": error}, 400)
        keys = ", ".join(expense.keys())
        placeholders = ", ".join([f":{key}" for key in expense])
        backup_database("expense-create")
        with db() as conn:
            cur = conn.execute(f"INSERT INTO expenses ({keys}) VALUES ({placeholders})", expense)
            saved = conn.execute("SELECT * FROM expenses WHERE id = ?", (cur.lastrowid,)).fetchone()
        return self.send_json(row_to_dict(saved), 201)

    def list_expenses(self, parsed):
        if not self.authorized():
            return self.send_json({"error": "Unauthorized"}, 401)
        query = parse_qs(parsed.query)
        month = query.get("month", [datetime.now().strftime("%Y-%m")])[0]
        start, end = month_bounds(month)
        with db() as conn:
            rows = conn.execute(
                "SELECT * FROM expenses WHERE date BETWEEN ? AND ? ORDER BY date ASC, id ASC",
                (start, end),
            ).fetchall()
        return self.send_json([row_to_dict(row) for row in rows])

    def update_expense(self, parsed):
        if not self.authorized():
            return self.send_json({"error": "Unauthorized"}, 401)
        expense_id = parsed.path.rsplit("/", 1)[-1]
        expense = normalize_expense(self.read_json())
        error = expense_error(expense)
        if error:
            return self.send_json({"error": error}, 400)
        assignments = ", ".join([f"{key}=:{key}" for key in expense])
        backup_database("expense-update")
        with db() as conn:
            cur = conn.execute(
                f"UPDATE expenses SET {assignments}, updatedAt=CURRENT_TIMESTAMP WHERE id=:id",
                {**expense, "id": expense_id},
            )
            if cur.rowcount == 0:
                return self.send_json({"error": "Δεν βρέθηκε το έξοδο."}, 404)
            saved = conn.execute("SELECT * FROM expenses WHERE id = ?", (expense_id,)).fetchone()
        return self.send_json(row_to_dict(saved))

    def expense_summary(self, parsed):
        if not self.authorized():
            return self.send_json({"error": "Unauthorized"}, 401)
        month = parse_qs(parsed.query).get("month", [datetime.now().strftime("%Y-%m")])[0]
        start, end = month_bounds(month)
        with db() as conn:
            revenue = conn.execute(
                "SELECT COALESCE(SUM(price), 0) AS total FROM bookings WHERE date BETWEEN ? AND ? AND status != 'Cancelled'",
                (start, end),
            ).fetchone()["total"]
            expenses = conn.execute(
                "SELECT COALESCE(SUM(amount), 0) AS total FROM expenses WHERE date BETWEEN ? AND ?",
                (start, end),
            ).fetchone()["total"]
            by_category = conn.execute(
                "SELECT category, COALESCE(SUM(amount), 0) AS total FROM expenses WHERE date BETWEEN ? AND ? GROUP BY category ORDER BY total DESC",
                (start, end),
            ).fetchall()
        return self.send_json({
            "revenue": revenue,
            "expenses": expenses,
            "net": revenue - expenses,
            "byCategory": [row_to_dict(row) for row in by_category],
        })

    def reports_monthly(self, parsed):
        if not self.authorized():
            return self.send_json({"error": "Unauthorized"}, 401)
        year, month, _, _, _ = parse_report_period(parse_qs(parsed.query))
        return self.send_json(fetch_monthly_report(year, month))

    def reports_monthly_export_xlsx(self, parsed):
        if not self.authorized():
            return self.send_json({"error": "Unauthorized"}, 401)
        if not OPENPYXL_AVAILABLE:
            return self.send_json({"error": "openpyxl is not installed on the server."}, 500)
        try:
            year, month, month_key, _, _ = parse_report_period(parse_qs(parsed.query))
            report = fetch_monthly_report(year, month)
            body = build_monthly_report_xlsx(report)
            filename = f"safewheels_report_{month_key}.xlsx"
            return self.send_file(
                body,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename,
            )
        except Exception as exc:
            return self.send_json({"error": f"Report export failed: {exc}"}, 500)

    def monthly_revenue(self, parsed):
        if not self.authorized():
            return self.send_json({"error": "Unauthorized"}, 401)
        query = parse_qs(parsed.query)
        try:
            year = int(query.get("year", [datetime.now(APP_TZ).year])[0])
        except ValueError:
            year = datetime.now(APP_TZ).year

        months = []
        season = {"cash": 0, "card": 0, "total": 0, "expenses": 0, "net": 0}
        with db() as conn:
            for month_num, month_name in SEASON_MONTHS:
                start = date(year, month_num, 1)
                next_month = (start.replace(day=28) + timedelta(days=4)).replace(day=1)
                end = next_month - timedelta(days=1)
                start_iso = start.isoformat()
                end_iso = end.isoformat()

                bookings = conn.execute(
                    """
                    SELECT date, pickupTime, route, customerName, price, paymentMethod
                    FROM bookings
                    WHERE date BETWEEN ? AND ? AND status != 'Cancelled'
                    ORDER BY date ASC, pickupTime ASC
                    """,
                    (start_iso, end_iso),
                ).fetchall()
                expenses = conn.execute(
                    """
                    SELECT date, description, category, amount, vehicle
                    FROM expenses
                    WHERE date BETWEEN ? AND ?
                    ORDER BY date ASC, id ASC
                    """,
                    (start_iso, end_iso),
                ).fetchall()

                entries = []
                cash_total = 0
                card_total = 0
                expense_total = 0
                for booking in bookings:
                    price = float(booking["price"] or 0)
                    payment_method = booking["paymentMethod"] or "Μετρητά"
                    is_cash = payment_method == "Μετρητά"
                    is_card = payment_method in ("Κάρτα", "Πίστωση")
                    is_credit = payment_method == "Πίστωση"
                    if is_card:
                        card_total += price
                    else:
                        cash_total += price
                    entries.append({
                        "type": "booking",
                        "date": booking["date"],
                        "time": booking["pickupTime"],
                        "route": booking["route"] or "",
                        "customer": booking["customerName"] or "",
                        "cash": price if is_cash else 0,
                        "card": price if is_card else 0,
                        "expenses": 0,
                        "description": "Πίστωση" if is_credit else "",
                    })
                for expense in expenses:
                    amount = float(expense["amount"] or 0)
                    expense_total += amount
                    description = expense["description"] or expense["category"] or ""
                    if expense["vehicle"]:
                        description = f"{description} · {expense['vehicle']}" if description else expense["vehicle"]
                    entries.append({
                        "type": "expense",
                        "date": expense["date"],
                        "time": "",
                        "route": "",
                        "customer": "",
                        "cash": 0,
                        "card": 0,
                        "expenses": amount,
                        "description": description,
                    })
                entries.sort(key=lambda item: (item["date"], item["time"] or "99:99", item["type"]))
                total = cash_total + card_total
                net = total - expense_total
                month_data = {
                    "month": month_num,
                    "name": month_name,
                    "cash": cash_total,
                    "card": card_total,
                    "total": total,
                    "expenses": expense_total,
                    "net": net,
                    "entries": entries,
                }
                months.append(month_data)
                for key in season:
                    season[key] += month_data[key]
        return self.send_json({"year": year, "months": months, "season": season})

    def option_table(self, kind):
        return {
            "vehicles": "vehicles",
            "drivers": "drivers",
            "sources": "booking_sources",
        }.get(kind)

    def options(self):
        if not self.authorized():
            return self.send_json({"error": "Unauthorized"}, 401)
        with db() as conn:
            return self.send_json({
                "vehicles": active_option_names(conn, "vehicles"),
                "drivers": active_option_names(conn, "drivers"),
                "bookingSources": active_option_names(conn, "booking_sources"),
                "optionDetails": {
                    "vehicles": active_option_rows(conn, "vehicles"),
                    "drivers": active_option_rows(conn, "drivers"),
                    "sources": active_option_rows(conn, "booking_sources"),
                },
                "paymentMethods": ["Μετρητά", "Κάρτα", "Πίστωση"],
                "taxStatuses": ["Καταχωρημένο", "Μη Καταχωρημένο"],
            })

    def create_option(self):
        if not self.authorized():
            return self.send_json({"error": "Unauthorized"}, 401)
        kind = self.path.rsplit("/", 1)[-1]
        table = self.option_table(kind)
        if not table:
            return self.send_json({"error": "Άγνωστη επιλογή."}, 404)
        name = str(self.read_json().get("name") or "").strip()
        if not name:
            return self.send_json({"error": "Το όνομα είναι υποχρεωτικό."}, 400)
        if kind == "sources":
            name = name.upper()
        backup_database(f"option-create-{kind}")
        with db() as conn:
            upsert_option(conn, table, name)
            saved = conn.execute(f"SELECT * FROM {table} WHERE name = ?", (name,)).fetchone()
        return self.send_json(row_to_dict(saved), 201)

    def update_option(self, parsed):
        if not self.authorized():
            return self.send_json({"error": "Unauthorized"}, 401)
        parts = parsed.path.strip("/").split("/")
        if len(parts) != 4:
            return self.send_json({"error": "Λάθος διεύθυνση επιλογής."}, 404)
        _, _, kind, item_id = parts
        table = self.option_table(kind)
        if not table:
            return self.send_json({"error": "Άγνωστη επιλογή."}, 404)
        name = str(self.read_json().get("name") or "").strip()
        if not name:
            return self.send_json({"error": "Το όνομα είναι υποχρεωτικό."}, 400)
        if kind == "sources":
            name = name.upper()
        backup_database(f"option-update-{kind}")
        with db() as conn:
            cur = conn.execute(
                f"UPDATE {table} SET name = ?, active = 1, updatedAt = CURRENT_TIMESTAMP WHERE id = ?",
                (name, item_id),
            )
            if cur.rowcount == 0:
                return self.send_json({"error": "Δεν βρέθηκε η επιλογή."}, 404)
            saved = conn.execute(f"SELECT * FROM {table} WHERE id = ?", (item_id,)).fetchone()
        return self.send_json(row_to_dict(saved))

    def delete_option(self, parsed):
        if not self.authorized():
            return self.send_json({"error": "Unauthorized"}, 401)
        parts = parsed.path.strip("/").split("/")
        if len(parts) != 4:
            return self.send_json({"error": "Λάθος διεύθυνση επιλογής."}, 404)
        _, _, kind, item_id = parts
        table = self.option_table(kind)
        if not table:
            return self.send_json({"error": "Άγνωστη επιλογή."}, 404)
        backup_database(f"option-delete-{kind}")
        with db() as conn:
            cur = conn.execute(
                f"UPDATE {table} SET active = 0, updatedAt = CURRENT_TIMESTAMP WHERE id = ?",
                (item_id,),
            )
        if cur.rowcount == 0:
            return self.send_json({"error": "Δεν βρέθηκε η επιλογή."}, 404)
        self.send_response(204)
        self.end_headers()

    def update_booking_tax_status(self, parsed):
        if not self.authorized():
            return self.send_json({"error": "Unauthorized"}, 401)
        booking_id = parsed.path.strip("/").split("/")[-2]
        tax_status = str(self.read_json().get("taxStatus") or "").strip()
        if tax_status not in TAX_STATUSES:
            return self.send_json({"error": "Λάθος κατάσταση ΑΑΔΕ."}, 400)
        backup_database("booking-tax-status")
        with db() as conn:
            cur = conn.execute(
                "UPDATE bookings SET taxStatus = ?, updatedAt = CURRENT_TIMESTAMP WHERE id = ?",
                (tax_status, booking_id),
            )
            if cur.rowcount == 0:
                return self.send_json({"error": "Δεν βρέθηκε η κράτηση."}, 404)
            saved = conn.execute("SELECT * FROM bookings WHERE id = ?", (booking_id,)).fetchone()
            audit_booking(conn, booking_id, "tax_status_updated")
        return self.send_json(row_to_dict(saved))

    def update_booking_status(self, parsed):
        if not self.authorized():
            return self.send_json({"error": "Unauthorized"}, 401)
        booking_id = parsed.path.strip("/").split("/")[-2]
        status = str(self.read_json().get("status") or "").strip()
        if status not in BOOKING_STATUSES:
            return self.send_json({"error": "Λάθος κατάσταση κράτησης."}, 400)
        backup_database("booking-status")
        with db() as conn:
            cur = conn.execute(
                "UPDATE bookings SET status = ?, updatedAt = CURRENT_TIMESTAMP WHERE id = ?",
                (status, booking_id),
            )
            if cur.rowcount == 0:
                return self.send_json({"error": "Δεν βρέθηκε η κράτηση."}, 404)
            saved = conn.execute("SELECT * FROM bookings WHERE id = ?", (booking_id,)).fetchone()
            audit_booking(conn, booking_id, "status_updated")
        return self.send_json(row_to_dict(saved))

    def admin_debug(self, parsed):
        if not self.authorized():
            return self.send_json({"error": "Unauthorized"}, 401)
        query = parse_qs(parsed.query)
        today = app_today().isoformat()
        start = normalize_date(query.get("start", [""])[0]) or today
        end = normalize_date(query.get("end", [""])[0]) or "9999-12-31"
        q = query.get("q", [""])[0].strip()
        tax = query.get("tax", [""])[0]
        driver = query.get("driver", [""])[0].strip()
        visible_where = ["date BETWEEN :start AND :end", "status NOT IN ('Completed', 'Cancelled')"]
        active_range_where = list(visible_where)
        params = {"start": start, "end": end}
        if q:
            visible_where.append("(CAST(id AS TEXT) LIKE :q OR customerName LIKE :q OR phone LIKE :q OR hotel LIKE :q OR date LIKE :q OR pickupTime LIKE :q OR flightNumber LIKE :q OR route LIKE :q)")
            params["q"] = f"%{q}%"
        if tax in TAX_STATUSES:
            visible_where.append("taxStatus = :tax")
            params["tax"] = tax
        if driver:
            visible_where.append("driver = :driver")
            params["driver"] = driver
        visible_sql = "SELECT COUNT(*) AS total FROM bookings WHERE " + " AND ".join(visible_where)
        active_range_sql = "SELECT COUNT(*) AS total FROM bookings WHERE " + " AND ".join(active_range_where)
        with db() as conn:
            db_rows = conn.execute("SELECT COUNT(*) AS total FROM bookings").fetchone()["total"]
            upcoming = conn.execute(
                "SELECT COUNT(*) AS total FROM bookings WHERE date >= ? AND status NOT IN ('Completed', 'Cancelled')",
                (today,),
            ).fetchone()["total"]
            completed = conn.execute("SELECT COUNT(*) AS total FROM bookings WHERE status = 'Completed'").fetchone()["total"]
            cancelled = conn.execute("SELECT COUNT(*) AS total FROM bookings WHERE status = 'Cancelled'").fetchone()["total"]
            active_in_range = conn.execute(active_range_sql, {"start": start, "end": end}).fetchone()["total"]
            visible = conn.execute(visible_sql, params).fetchone()["total"]
            audit_rows = conn.execute("SELECT COUNT(*) AS total FROM booking_audit_log").fetchone()["total"]
            matches = []
            if q:
                match_where = ["(CAST(id AS TEXT) LIKE :q OR customerName LIKE :q OR phone LIKE :q OR hotel LIKE :q OR date LIKE :q OR pickupTime LIKE :q OR flightNumber LIKE :q OR route LIKE :q)"]
                match_params = {"q": f"%{q}%"}
                if tax in TAX_STATUSES:
                    match_where.append("taxStatus = :tax")
                    match_params["tax"] = tax
                if driver:
                    match_where.append("driver = :driver")
                    match_params["driver"] = driver
                matches = [
                    row_to_dict(row)
                    for row in conn.execute(
                        "SELECT id, date, pickupTime, customerName, route, status, driver, taxStatus, updatedAt FROM bookings WHERE "
                        + " AND ".join(match_where)
                        + " ORDER BY date DESC, pickupTime DESC LIMIT 20",
                        match_params,
                    ).fetchall()
                ]
        return self.send_json({
            "totalBookings": db_rows,
            "upcoming": upcoming,
            "completed": completed,
            "cancelled": cancelled,
            "activeInCurrentRange": active_in_range,
            "visibleWithCurrentFilters": visible,
            "hiddenByFilters": max(active_in_range - visible, 0),
            "databaseRowsCount": db_rows,
            "auditRowsCount": audit_rows,
            "matches": matches,
            "range": {"start": start, "end": end},
        })

    def admin_create_backup(self):
        if not self.authorized():
            return self.send_json({"error": "Unauthorized"}, 401)
        try:
            backup_path = create_live_backup("live")
            body = backup_path.read_bytes()
            return self.send_file(
                body,
                "application/x-sqlite3",
                backup_path.name,
            )
        except Exception as exc:
            return self.send_json({"error": f"Backup failed: {exc}"}, 500)

    def admin_list_backups(self):
        if not self.authorized():
            return self.send_json({"error": "Unauthorized"}, 401)
        return self.send_json({"backups": list_live_backups()})

    def admin_download_backup(self, filename):
        if not self.authorized():
            return self.send_json({"error": "Unauthorized"}, 401)
        if not is_allowed_live_backup_name(filename):
            return self.send_json({"error": "Invalid backup filename."}, 400)
        backup_path = LIVE_BACKUP_DIR / Path(filename).name
        if not backup_path.exists():
            return self.send_json({"error": "Backup not found."}, 404)
        body = backup_path.read_bytes()
        return self.send_file(body, "application/x-sqlite3", backup_path.name)

    def admin_export_bookings_xlsx(self):
        if not self.authorized():
            return self.send_json({"error": "Unauthorized"}, 401)
        if not OPENPYXL_AVAILABLE:
            return self.send_json({"error": "openpyxl is not installed on the server."}, 500)
        try:
            body = build_bookings_xlsx()
            filename = f"safewheels_bookings_export_{datetime.now(APP_TZ).strftime('%Y-%m-%d_%H-%M')}.xlsx"
            return self.send_file(
                body,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename,
            )
        except Exception as exc:
            return self.send_json({"error": f"Export failed: {exc}"}, 500)

    def admin_restore_backup(self):
        if not self.authorized():
            return self.send_json({"error": "Unauthorized"}, 401)
        content_type = self.headers.get("Content-Type", "")
        if "multipart/form-data" not in content_type:
            return self.send_json({"error": "Expected multipart/form-data upload."}, 400)
        length = int(self.headers.get("Content-Length") or 0)
        if length <= 0:
            return self.send_json({"error": "Empty upload."}, 400)
        body = self.rfile.read(length)
        try:
            if "multipart/form-data" in content_type:
                file_bytes = parse_multipart_upload(content_type, body)
            elif content_type == "application/octet-stream":
                file_bytes = body
            else:
                return self.send_json({"error": "Expected multipart/form-data upload."}, 400)
        except ValueError as exc:
            return self.send_json({"error": str(exc)}, 400)
        LIVE_BACKUP_DIR.mkdir(parents=True, exist_ok=True)
        temp_path = LIVE_BACKUP_DIR / f"restore_upload_{live_backup_timestamp()}.sqlite"
        try:
            with temp_path.open("wb") as handle:
                handle.write(file_bytes)
            ok, details, emergency_name = restore_sqlite_backup(temp_path)
            if not ok:
                return self.send_json(
                    {
                        "error": details,
                        "emergencyBackup": emergency_name,
                        "restored": False,
                    },
                    400,
                )
            return self.send_json(
                {
                    "restored": True,
                    "emergencyBackup": emergency_name,
                    "details": details,
                }
            )
        except Exception as exc:
            return self.send_json({"error": f"Restore failed: {exc}", "restored": False}, 500)
        finally:
            temp_path.unlink(missing_ok=True)

    def delete_row(self, table, parsed):
        if not self.authorized():
            return self.send_json({"error": "Unauthorized"}, 401)
        item_id = parsed.path.rsplit("/", 1)[-1]
        backup_database(f"{table}-delete")
        with db() as conn:
            if table == "bookings":
                # Store full snapshot in audit log before permanent deletion
                audit_booking(conn, item_id, "deleted")
            cur = conn.execute(f"DELETE FROM {table} WHERE id = ?", (item_id,))
        if cur.rowcount == 0:
            return self.send_json({"error": "Δεν βρέθηκε η εγγραφή."}, 404)
        self.send_response(204)
        self.end_headers()

    def authorized(self):
        return self.headers.get("x-safewheels-session") in SESSIONS

    def read_json(self):
        length = int(self.headers.get("content-length") or 0)
        if not length:
            return {}
        return json.loads(self.rfile.read(length).decode("utf-8"))

    def send_json(self, data, status=200):
        body = json.dumps(data, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def send_file(self, body, content_type, filename):
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)


if __name__ == "__main__":
    init_db()
    server = ThreadingHTTPServer(("0.0.0.0", PORT), Handler)
    print(f"SafeWheels booking app running at http://localhost:{PORT}")
    server.serve_forever()
